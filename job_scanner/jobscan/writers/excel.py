"""The workbook: every scored posting, plus the two sheets that say whether
the numbers can be trusted.

Run status is not an appendix. A scan that reaches four employers and finds
nothing is indistinguishable, in the postings sheet alone, from a scan whose
four employers all returned HTTP 403 -- so status, fetched, kept and the error
text sit beside each other on their own sheet, and Run info records what the
run was configured with.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SHORTLIST_FILL = PatternFill("solid", fgColor="E2EFDA")

STATUS_FILLS = {
    "ok": PatternFill("solid", fgColor="E2EFDA"),
    "empty": PatternFill("solid", fgColor="FFF2CC"),
    "error": PatternFill("solid", fgColor="FCE4E4"),
    "blocked": PatternFill("solid", fgColor="FCE4E4"),
    "skipped": PatternFill("solid", fgColor="F2F2F2"),
}

POSTING_COLUMNS = [
    ("source", 14), ("title", 52), ("grade", 20), ("department", 26),
    ("location", 26), ("country", 18), ("posted_at", 12), ("closing_at", 12),
    ("age_days", 10), ("score", 8), ("shortlisted", 12), ("matched", 30),
    ("why", 70), ("url", 60),
]

STATUS_COLUMNS = [
    ("source", 16), ("name", 40), ("platform", 16), ("status", 10),
    ("fetched", 9), ("kept", 8), ("verified", 15), ("endpoint", 46),
    ("error", 80), ("notes", 80), ("ms", 8),
]


def _write_sheet(sheet, columns, rows) -> None:
    for index, (name, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index, row in enumerate(rows, start=2):
        for col_index, (name, _) in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=col_index, value=row.get(name, ""))
    sheet.freeze_panes = "A2"


def write_xlsx(path, postings, run_log, config, dropped_stale=0, unknown_dates=0) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    book = Workbook()

    sheet = book.active
    sheet.title = "Postings"
    ordered = sorted(postings, key=lambda p: (-p.score, p.source_key, p.title))
    _write_sheet(sheet, POSTING_COLUMNS, [p.as_row() for p in ordered])
    shortlist_column = [name for name, _ in POSTING_COLUMNS].index("shortlisted") + 1
    for row_index, posting in enumerate(ordered, start=2):
        if posting.shortlisted:
            for col_index in range(1, len(POSTING_COLUMNS) + 1):
                sheet.cell(row=row_index, column=col_index).fill = SHORTLIST_FILL
    sheet.cell(row=1, column=shortlist_column).alignment = Alignment(horizontal="center")

    status = book.create_sheet("Run status")
    records = run_log.records
    _write_sheet(status, STATUS_COLUMNS, [r.as_row() for r in records])
    status_column = [name for name, _ in STATUS_COLUMNS].index("status") + 1
    for row_index, record in enumerate(records, start=2):
        fill = STATUS_FILLS.get(record.status)
        if fill:
            status.cell(row=row_index, column=status_column).fill = fill

    info = book.create_sheet("Run info")
    totals = run_log.totals()
    facts = [
        ("generated_at", _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")),
        ("config_file", str(config.path)),
        ("profile_target", config.profile.get("target", "")),
        ("profile_fingerprint", config.fingerprint()),
        ("max_age_days", config.max_age_days),
        ("shortlist_min_score", config.shortlist_min_score),
        ("sources_in_config", len(config.sources)),
        ("sources_in_report", totals["sources"]),
        ("sources_attempted", totals["attempted"]),
        ("sources_ok", totals["ok"]),
        ("sources_empty", totals["empty"]),
        ("sources_error", totals["error"]),
        ("sources_blocked", totals["blocked"]),
        ("sources_skipped", totals["skipped"]),
        ("postings_fetched", totals["fetched"]),
        ("postings_kept", totals["kept"]),
        ("postings_shortlisted", sum(1 for p in postings if p.shortlisted)),
        ("dropped_older_than_max_age", dropped_stale),
        ("kept_with_unknown_posting_date", unknown_dates),
    ]
    info.column_dimensions["A"].width = 34
    info.column_dimensions["B"].width = 76
    for row_index, (key, value) in enumerate(facts, start=1):
        info.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        info.cell(row=row_index, column=2, value=value)

    book.save(target)
    return target
