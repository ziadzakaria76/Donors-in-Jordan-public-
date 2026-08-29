"""Excel export.

Two openpyxl footguns are handled here: fills need BARE hex (openpyxl rejects
#RRGGBB), and auto_filter over an empty sheet is a classic crash -- a run with
zero tenders must still produce a valid workbook.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .common import fmt_value

HEADER_FILL = PatternFill("solid", fgColor="1B4F72")     # bare hex, not #1B4F72
NEW_FILL = PatternFill("solid", fgColor="FFF3CD")

COLUMNS = [
    ("NEW", lambda t: "NEW" if t.is_new else ""),
    ("Score", lambda t: t.score),
    ("Title", lambda t: t.title),
    ("Portal", lambda t: t.portal),
    ("Notice type", lambda t: t.notice_type or ""),
    ("Posted", lambda t: t.posted_date.isoformat() if t.posted_date else ""),
    ("Closes", lambda t: t.closing_date.isoformat() if t.closing_date else "not published"),
    ("Value", fmt_value),
    ("Currency", lambda t: t.raw_currency or ""),
    ("Sector", lambda t: t.sector or ""),
    ("Eligibility", lambda t: t.eligibility or ""),
    ("Contact", lambda t: t.contact or ""),
    ("Language", lambda t: t.language or ""),
    ("Delivery country", lambda t: t.delivery_country or ""),
    ("Syria link", lambda t: t.syria_link_type),
    ("Sanctions flag", lambda t: "; ".join(h["matched_name"] for h in t.screening)),
    ("Flags", lambda t: "; ".join(t.flags)),
    ("URL", lambda t: t.url or ""),
]


def write_xlsx(result, path: Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Tenders"
    _sheet(ws, result.tenders)

    _sheet(wb.create_sheet("Excluded from scope"), result.excluded)

    health = wb.create_sheet("Run diagnostics")
    health.append(["Portal", "Available", "Fetched", "Kept", "Layer", "Quality", "Error/skip", "URL"])
    for portal in result.portals:
        health.append([portal.label, "yes" if portal.available else "NO", portal.stats.seen,
                       len(portal.tenders), portal.layer or "", portal.quality or "",
                       portal.skipped_reason or portal.error or "", portal.url])
    health.append([])
    health.append(["Classification split"])
    for key, count in result.counts.items():
        health.append([key, count])
    health.append([])
    health.append(["Sanctions lists (triage aid, never legal clearance)"])
    for entry in result.screening_status:
        health.append([entry["list"], f"fetched {entry['fetched']}", f"{entry['names']} names",
                       entry.get("error") or ""])
    if result.screening_error:
        health.append(["SCREENING ERROR", result.screening_error])

    wb.save(path)
    return path


def _sheet(ws, tenders) -> None:
    ws.append([name for name, _ in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for tender in tenders:
        ws.append([getter(tender) for _, getter in COLUMNS])
        if tender.is_new:
            for cell in ws[ws.max_row]:
                cell.fill = NEW_FILL

    widths = [8, 8, 60, 12, 16, 12, 12, 22, 10, 20, 28, 26, 10, 16, 20, 24, 28, 46]
    for idx, width in enumerate(widths[:len(COLUMNS)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # auto_filter over an empty sheet crashes on open; only add it with rows.
    if tenders:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    ws.freeze_panes = "A2"
