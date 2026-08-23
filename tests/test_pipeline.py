"""End-to-end pipeline behaviour, offline."""

from __future__ import annotations

from datetime import date, timedelta

import pytest

from syria_monitor.fetch import Fetcher
from syria_monitor.models import Tender
from syria_monitor.pipeline import run as run_pipeline
from syria_monitor.portals import REGISTRY
from syria_monitor.portals.base import BasePortal
from syria_monitor.report import render_email, write_docx, write_json, write_xlsx
from syria_monitor.scoring import score_batch
from syria_monitor.state import SeenStore

TODAY = date(2026, 8, 23)


def make_portal(name, records, fails=False):
    class Fake(BasePortal):
        pass
    Fake.name = name
    Fake.label = name.upper()
    Fake.url = f"https://{name}.example/notices"

    def fetch_tenders(self):
        if fails:
            raise RuntimeError("boom")
        return records
    Fake.fetch_tenders = fetch_tenders
    return Fake


@pytest.fixture
def registry(monkeypatch):
    original = dict(REGISTRY)
    monkeypatch.setattr("syria_monitor.pipeline.REGISTRY", REGISTRY)
    yield REGISTRY
    REGISTRY.clear()
    REGISTRY.update(original)


def syria_record(**kw):
    base = {"id": "1", "title": "Rehabilitation of the Aleppo water network",
            "project_ctry_name": "Syrian Arab Republic", "closing_date": "2026-09-30",
            "_safe_text_fields": ["title"]}
    base.update(kw)
    return base


# ---------------------------------------------------------------- empty inputs
def test_empty_run_is_not_an_error(config, registry):
    registry.clear()
    registry["empty"] = make_portal("empty", [])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["empty"])
    assert result.tenders == []
    assert "no new opportunities" in result.subject()
    assert "all 1 portals OK" in result.subject()


def test_tender_with_every_optional_field_none_survives(config, registry):
    registry.clear()
    registry["sparse"] = make_portal("sparse", [{"id": "x", "title": "Works in Damascus and Aleppo",
                                                 "_safe_text_fields": ["title"]}])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["sparse"])
    assert len(result.tenders) == 1
    tender = result.tenders[0]
    assert tender.closing_date is None
    assert "deadline_not_published" in tender.flags


def test_scoring_with_zero_lexicon_matches_does_not_divide_by_zero(profile):
    tenders = [Tender(id=str(i), title="zzz qqq", portal="p") for i in range(3)]
    score_batch(tenders, profile, TODAY)
    assert all(t.score >= 0 for t in tenders)


def test_scoring_drops_components_that_carry_no_information(profile):
    """With one link type and no values, those components award everyone the
    same points; the surviving weights renormalise to 100."""
    tenders = [
        Tender(id="1", title="Water supply rehabilitation", portal="p",
               closing_date=TODAY + timedelta(days=10), syria_link_type="inside_syria"),
        Tender(id="2", title="Vehicle supply", portal="p",
               closing_date=TODAY + timedelta(days=200), syria_link_type="inside_syria"),
    ]
    score_batch(tenders, profile, TODAY)
    weights = [e for e in tenders[0].match_evidence if e.startswith("weights:")][0]
    assert "link_type" not in weights          # identical for every tender
    assert "value" not in weights              # no values published
    assert abs(sum(float(p.split("=")[1]) for p in weights[len("weights:"):].split(",")) - 100) < 0.5


# ------------------------------------------------------------------- deadlines
def test_deadline_today_is_kept_and_yesterday_is_dropped(config, registry):
    registry.clear()
    registry["dl"] = make_portal("dl", [
        syria_record(id="today", closing_date=TODAY.isoformat()),
        syria_record(id="yesterday", closing_date=(TODAY - timedelta(days=1)).isoformat()),
        syria_record(id="none", closing_date=None),
    ])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["dl"])
    kept = {t.id for t in result.tenders}
    assert "today" in kept and "none" in kept and "yesterday" not in kept
    assert result.expired_dropped == 1


# ------------------------------------------------------------------ duplicates
def test_duplicates_collapse_across_portals(config, registry):
    registry.clear()
    registry["a"] = make_portal("a", [syria_record(id="a1")])
    registry["b"] = make_portal("b", [syria_record(id="b1")])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["a", "b"])
    assert len(result.tenders) == 1
    assert result.duplicates_collapsed == 1
    assert any(f.startswith("also_on:") for f in result.tenders[0].flags)


# --------------------------------------------------------------- portal health
def test_one_failing_portal_does_not_abort_the_run(config, registry):
    registry.clear()
    registry["good"] = make_portal("good", [syria_record()])
    registry["bad"] = make_portal("bad", [], fails=True)
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["good", "bad"])
    assert len(result.tenders) == 1
    assert len(result.failed) == 1
    assert "WARNING: 1 portal(s) down" in result.subject()
    assert "bad.example" in result.failed[0].url


def test_all_portals_down_says_action_needed(config, registry):
    registry.clear()
    registry["bad"] = make_portal("bad", [], fails=True)
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["bad"])
    assert result.subject().startswith("ACTION NEEDED: all portals unreachable")


def test_a_quiet_day_and_a_broken_monitor_have_different_subjects(config, registry):
    registry.clear()
    registry["quiet"] = make_portal("quiet", [])
    quiet = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["quiet"]).subject()
    registry.clear()
    registry["broken"] = make_portal("broken", [], fails=True)
    broken = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["broken"]).subject()
    assert quiet != broken
    assert "OK" in quiet and "ACTION NEEDED" in broken


# ----------------------------------------------------------------------- scope
def test_out_of_scope_categories_are_counted_not_silently_dropped(config, registry):
    registry.clear()
    registry["mix"] = make_portal("mix", [
        syria_record(id="inside"),
        {"id": "refugee", "title": "Education for Syrian refugee children, Mafraq",
         "country": "Jordan", "closing_date": "2026-09-30", "_safe_text_fields": ["title"]},
        {"id": "hub", "title": "Whole of Syria WASH coordination, duty station Gaziantep",
         "place_of_performance_country": "TR", "closing_date": "2026-09-30",
         "_safe_text_fields": ["title"]},
    ])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["mix"])
    assert [t.id for t in result.tenders] == ["inside"]
    assert result.counts["refugee_hosting_only"] == 1
    assert result.counts["cross_border_hub"] == 1
    # Both related-but-out-of-scope categories are kept line by line for audit,
    # not reduced to a count -- a misclassification has to be visible to be
    # correctable, and widening scope later is a config change.
    assert sorted(t.id for t in result.excluded) == ["hub", "refugee"]
    assert {t.syria_link_type for t in result.excluded} == {"cross_border_hub",
                                                            "refugee_hosting_only"}


# --------------------------------------------------------------------- outputs
@pytest.fixture
def arabic_result(config, registry):
    registry.clear()
    registry["ar"] = make_portal("ar", [
        {"id": "ar1", "title": "إعادة تأهيل شبكة المياه — حلب",
         "description": "الجمهورية العربية السورية", "closing_date": "2026-09-30",
         "project_ctry_name": "Syrian Arab Republic", "_safe_text_fields": ["title"]},
        syria_record(id="en1"),
    ])
    return run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["ar"])


def test_arabic_survives_json_and_excel(arabic_result, config, tmp_path):
    json_path = write_json(arabic_result, tmp_path / "r.json", config.profile)
    assert "إعادة تأهيل" in json_path.read_text(encoding="utf-8")

    xlsx_path = write_xlsx(arabic_result, tmp_path / "r.xlsx")
    from openpyxl import load_workbook
    titles = [row[2] for row in load_workbook(xlsx_path).active.iter_rows(values_only=True)]
    assert any(t and "إعادة تأهيل" in t for t in titles)


def test_rtl_survives_the_word_writer(arabic_result, tmp_path):
    """True RTL is w:bidi on the paragraph and w:rtl on the runs -- right
    alignment alone is not RTL."""
    import zipfile
    path = write_docx(arabic_result, tmp_path / "r.docx")
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8")
    assert "إعادة تأهيل" in xml
    assert "<w:bidi" in xml
    assert "<w:rtl" in xml


def test_zero_row_workbook_opens(config, registry, tmp_path):
    """auto_filter on an empty sheet is a classic crash."""
    registry.clear()
    registry["empty"] = make_portal("empty", [])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["empty"])
    path = write_xlsx(result, tmp_path / "empty.xlsx")
    from openpyxl import load_workbook
    workbook = load_workbook(path)
    assert workbook.active.max_row == 1
    assert workbook.active.auto_filter.ref is None


def test_word_report_writes_with_zero_tenders(config, registry, tmp_path):
    registry.clear()
    registry["empty"] = make_portal("empty", [])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["empty"])
    assert write_docx(result, tmp_path / "empty.docx").exists()


def test_report_states_the_screening_disclaimer(arabic_result, tmp_path):
    import zipfile
    xml = zipfile.ZipFile(write_docx(arabic_result, tmp_path / "d.docx")).read(
        "word/document.xml").decode("utf-8")
    assert "triage aid, never legal clearance" in xml


# -------------------------------------------------------------------- delivery
def test_email_overflow_drops_attachments_rather_than_the_email(tmp_path):
    from syria_monitor.delivery import GraphMailer
    big = tmp_path / "big.xlsx"
    big.write_bytes(b"0" * (4 * 1024 * 1024))
    message, notes = GraphMailer.build_message("s", "<p>body</p>", ["a@b.c"], [], [big])
    assert message["message"]["attachments"] == []
    assert notes and "attachments omitted" in notes[0]
    assert "Attachments omitted" in message["message"]["body"]["content"]


def test_delivery_without_credentials_fails_clearly(monkeypatch):
    from syria_monitor.delivery import GraphMailer, MailError
    mailer = GraphMailer()
    with pytest.raises(MailError) as excinfo:
        mailer.send("s", "<p>x</p>", ["a@b.c"])
    assert "GRAPH_TENANT_ID" in str(excinfo.value)


def test_delivery_without_recipients_fails_clearly():
    from syria_monitor.delivery import GraphMailer, MailError
    with pytest.raises(MailError) as excinfo:
        GraphMailer().send("s", "<p>x</p>", [])
    assert "REPORT_TO" in str(excinfo.value)


def test_email_body_renders_without_tenders(config, registry):
    registry.clear()
    registry["empty"] = make_portal("empty", [])
    result = run_pipeline(config, fetcher=Fetcher(), today=TODAY, portals=["empty"])
    html = render_email(result)
    assert "Syria tenders" in html and "triage aid" in html


# ----------------------------------------------------------------------- state
def test_diagnostics_do_not_write_to_the_seen_database(tmp_path):
    """A --self-test that writes fixture ids into the real database makes the
    next real run report nothing, which looks exactly like a broken monitor."""
    db = tmp_path / "seen.db"
    tenders = [Tender(id="fixture-1", title="t", portal="p")]

    readonly = SeenStore(db, read_only=True)
    readonly.mark_new(tenders)
    assert readonly.record(tenders) == 0
    readonly.close()

    live = SeenStore(db)
    live.mark_new(tenders)
    assert tenders[0].is_new is True, "the fixture run must not have marked this as seen"
    live.close()


def test_new_marking_flips_after_a_real_run(tmp_path):
    db = tmp_path / "seen.db"
    tenders = [Tender(id="1", title="t", portal="p")]
    store = SeenStore(db)
    store.mark_new(tenders)
    assert tenders[0].is_new is True
    store.record(tenders)
    store.mark_new(tenders)
    assert tenders[0].is_new is False
    store.close()
